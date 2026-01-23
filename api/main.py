"""FastAPI application for Epitome frontend demo."""
import asyncio
import json
import shutil
import sys
import uuid
from datetime import datetime
from pathlib import Path
from typing import Optional

# Load environment variables from .env file BEFORE importing agents
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # python-dotenv not installed, continue without it

from fastapi import FastAPI, File, Form, HTTPException, UploadFile, Depends, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from sqlalchemy.ext.asyncio import AsyncSession

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from api.progress import ProgressEvent, progress_manager
from api.database import get_db, AsyncSessionLocal
from api.services import (
    create_project_from_generation,
    get_project_for_frontend,
    get_project_as_generator_data,
    update_crew_rsvp,
    search_crew_members,
    update_call_sheet,
    update_project,
    update_location,
)
from api.services.chat_service import process_chat_message
from agents.production_workbook_generator import run_tool, EpitomeWorkbookGenerator

# Directory setup
BASE_DIR = Path(__file__).parent.parent
OUTPUT_DIR = BASE_DIR / "output"
STATIC_DIR = BASE_DIR / "static"

OUTPUT_DIR.mkdir(exist_ok=True)

# Create FastAPI app
app = FastAPI(
    title="Epitome Production Workbook Generator",
    description="Generate production workbooks from natural language prompts",
    version="1.0.0"
)

# CORS for local development
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Mount static files (for Vite-built frontend assets)
# IMPORTANT: Mount more specific paths first, then general ones
if STATIC_DIR.exists():
    # Serve assets directory (JS, CSS from Vite build)
    assets_dir = STATIC_DIR / "assets"
    if assets_dir.exists():
        app.mount("/assets", StaticFiles(directory=str(assets_dir)), name="assets")
    
    # Serve other static files (favicon, robots.txt, etc.) from root of static
    # But exclude index.html and assets/ since they're handled separately
    # Note: FastAPI StaticFiles will serve files from the directory root
    # We'll handle index.html via the route handler instead


@app.get("/", response_class=HTMLResponse)
async def serve_index():
    """Serve the main HTML page from Vite build."""
    index_path = STATIC_DIR / "index.html"
    if not index_path.exists():
        return HTMLResponse(
            content="<h1>Epitome API</h1><p>Frontend not found. Run ./sync_frontend.sh to build the frontend.</p>",
            status_code=200
        )
    return FileResponse(str(index_path))


# Note: Catch-all route moved to end of file to avoid interfering with other routes


@app.post("/api/generate")
async def generate_workbook(
    prompt: Optional[str] = Form(None),
    file: Optional[UploadFile] = File(None)
):
    """
    Start workbook generation. Returns a job_id for progress tracking.

    - Accepts multipart form data
    - prompt: Optional natural language request (can be empty if file is provided)
    - file: Optional CSV/PDF/TXT/Excel file for context
    - At least one of prompt or file must be provided
    """
    # Validate that at least one of prompt or file is provided
    prompt_value = prompt.strip() if prompt else ""
    if not prompt_value and not file:
        raise HTTPException(
            status_code=400,
            detail="Either a prompt or a file must be provided"
        )
    
    loop = asyncio.get_event_loop()
    job_id = progress_manager.create_job(loop)

    # Read uploaded file if provided
    file_content = None
    if file:
        content = await file.read()
        filename_lower = file.filename.lower()

        # Handle different file types
        if filename_lower.endswith('.csv') or filename_lower.endswith('.txt'):
            file_content = content.decode('utf-8', errors='ignore')

        elif filename_lower.endswith('.pdf'):
            # Extract text from PDF
            try:
                import io
                import PyPDF2

                pdf_file = io.BytesIO(content)
                pdf_reader = PyPDF2.PdfReader(pdf_file)

                # Extract text from all pages
                extracted_text = []
                for page_num, page in enumerate(pdf_reader.pages, 1):
                    page_text = page.extract_text()
                    if page_text.strip():
                        extracted_text.append(f"--- Page {page_num} ---\n{page_text}")

                if extracted_text:
                    file_content = f"[PDF file: {file.filename}]\n\n" + "\n\n".join(extracted_text)
                else:
                    file_content = f"[PDF file: {file.filename} - No extractable text found]"

            except Exception as e:
                file_content = f"[PDF file: {file.filename} - Error extracting text: {str(e)}]"

        elif filename_lower.endswith('.xlsx') or filename_lower.endswith('.xls'):
            # Extract text from Excel files
            try:
                import io
                import openpyxl

                excel_file = io.BytesIO(content)
                workbook = openpyxl.load_workbook(excel_file, data_only=True)

                extracted_sheets = []
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    sheet_rows = []

                    for row in sheet.iter_rows(values_only=True):
                        # Filter out empty rows
                        row_values = [str(cell) if cell is not None else "" for cell in row]
                        if any(v.strip() for v in row_values):
                            sheet_rows.append(" | ".join(row_values))

                    if sheet_rows:
                        extracted_sheets.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(sheet_rows))

                if extracted_sheets:
                    file_content = f"[Excel file: {file.filename}]\n\n" + "\n\n".join(extracted_sheets)
                else:
                    file_content = f"[Excel file: {file.filename} - No data found]"

            except Exception as e:
                file_content = f"[Excel file: {file.filename} - Error extracting data: {str(e)}]"

        else:
            # Try to decode as text for other file types
            try:
                file_content = content.decode('utf-8', errors='ignore')
            except Exception:
                file_content = f"[File: {file.filename} - Could not read content]"

    # Run generation in background
    # Use empty string if prompt is None/empty
    prompt_to_use = prompt_value if prompt_value else ""
    asyncio.create_task(
        run_generation_task(job_id, prompt_to_use, file_content, loop)
    )

    return {"job_id": job_id, "status": "started"}


async def run_generation_task(
    job_id: str,
    prompt: str,
    file_content: Optional[str],
    loop: asyncio.AbstractEventLoop
):
    """Background task that runs the generation pipeline."""

    def progress_callback(stage_id: str, percent: int, message: str):
        """Callback to emit progress events."""
        progress_manager.emit_progress(job_id, stage_id, percent, message)

    try:
        # Run the synchronous run_tool in a thread pool
        result = await loop.run_in_executor(
            None,
            lambda: run_tool(
                prompt=prompt,
                attached_file_content=file_content,
                enrich=True,
                progress_callback=progress_callback
            )
        )

        # Move workbook to output directory with unique name
        progress_callback("saving_file", 92, "Saving workbook file...")
        original_path = Path(result['workbook_path'])
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        new_filename = f"Epitome_Workbook_{job_id[:8]}_{timestamp}.xlsx"
        new_path = OUTPUT_DIR / new_filename

        if original_path.exists():
            shutil.move(str(original_path), str(new_path))

        result['workbook_path'] = str(new_path)
        result['download_filename'] = new_filename

        # Save to database
        progress_callback("saving_database", 95, "Saving to database...")
        project_id = None
        try:
            enriched_data = result.get('data', {})
            # Debug: Log the data being saved
            print(f"[DB DEBUG] Saving to database...")
            print(f"[DB DEBUG] production_info: {enriched_data.get('production_info', {})}")
            crew_list = enriched_data.get('crew_list', [])
            print(f"[DB DEBUG] crew_list count: {len(crew_list)}")
            if crew_list:
                print(f"[DB DEBUG] First crew member: {crew_list[0]}")
            print(f"[DB DEBUG] schedule_days: {enriched_data.get('schedule_days', [])}")

            async with AsyncSessionLocal() as db:
                project_id = await create_project_from_generation(db, enriched_data)
                result['project_id'] = project_id
                print(f"[DB DEBUG] Project saved with ID: {project_id}")
        except Exception as db_error:
            # Log but don't fail - workbook was still generated
            error_type = type(db_error).__name__
            error_msg = str(db_error)
            print(f"Warning: Database save failed: {error_type}: {error_msg}")
            if "nodename" in error_msg.lower() or "servname" in error_msg.lower():
                print("  This is a DNS resolution error. Check your DATABASE_URL in .env file.")
                print("  The database hostname cannot be resolved. Verify:")
                print("  - DATABASE_URL is correctly formatted")
                print("  - Network connectivity to Supabase")
                print("  - DNS resolution is working")
            elif "connection" in error_msg.lower():
                print("  This is a connection error. Check:")
                print("  - DATABASE_URL is correct")
                print("  - Supabase database is accessible")
                print("  - Firewall/network settings")

        # Store enriched data in result for fallback when database save fails
        if not project_id and enriched_data:
            result['enriched_data'] = enriched_data
            result['job_id'] = job_id  # Store job_id for fallback endpoint
            print(f"[DB DEBUG] Stored enriched data in result (fallback mode, job_id: {job_id})")
        
        progress_manager.set_result(job_id, result)

        # Final ready stage
        progress_callback("ready", 99, "Your call sheet is ready!")

        # Send download ready event with project_id and job_id (for fallback)
        download_event = ProgressEvent(
            "download_ready", 100,
            json.dumps({"filename": new_filename, "project_id": project_id, "job_id": job_id if not project_id else None})
        )
        queue = progress_manager.get_queue(job_id)
        if queue:
            await queue.put(download_event)

    except Exception as e:
        # Emit error event
        error_event = ProgressEvent("error", -1, f"Error: {str(e)}")
        queue = progress_manager.get_queue(job_id)
        if queue:
            await queue.put(error_event)
        progress_manager.set_result(job_id, {"error": str(e)})


@app.get("/api/progress/{job_id}")
async def stream_progress(job_id: str):
    """
    SSE endpoint for real-time progress updates.

    Client connects and receives events like:
    data: {"stage_id": "extracting_data", "percent": 30, "message": "..."}
    """
    queue = progress_manager.get_queue(job_id)
    if not queue:
        raise HTTPException(status_code=404, detail="Job not found")

    async def event_generator():
        try:
            # Send immediate "connected" event to establish the stream
            connected_event = ProgressEvent("connected", 0, "Connected to progress stream")
            yield connected_event.to_sse()

            while True:
                try:
                    # Wait for next event with shorter timeout for more frequent keepalives
                    event = await asyncio.wait_for(queue.get(), timeout=30.0)
                    yield event.to_sse()

                    # Check if error or download ready (don't break on 'complete' - wait for download_ready)
                    if event.stage_id in ("error", "download_ready"):
                        break

                except asyncio.TimeoutError:
                    # Send keepalive comment to prevent proxy timeout
                    yield ": keepalive\n\n"

        except asyncio.CancelledError:
            pass
        finally:
            # Clean up after stream ends
            progress_manager.cleanup_job(job_id)

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache, no-transform",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
            "Content-Type": "text/event-stream",
        }
    )


@app.get("/api/download/{filename}")
async def download_workbook(filename: str):
    """Download a generated workbook."""
    # Security: sanitize filename
    safe_filename = Path(filename).name
    file_path = OUTPUT_DIR / safe_filename

    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")

    # Ensure file is within output directory
    try:
        file_path.resolve().relative_to(OUTPUT_DIR.resolve())
    except ValueError:
        raise HTTPException(status_code=403, detail="Access denied")

    return FileResponse(
        path=str(file_path),
        filename=safe_filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.get("/api/result/{job_id}")
async def get_result(job_id: str):
    """Get the result data for a completed job."""
    result = progress_manager.get_result(job_id)
    if not result:
        raise HTTPException(status_code=404, detail="Result not found")

    return {
        "status": "error" if "error" in result else "complete",
        "data": result.get("data"),
        "download_filename": result.get("download_filename"),
        "project_id": result.get("project_id"),
        "error": result.get("error")
    }


@app.get("/api/generation/{job_id}/data")
async def get_generation_data(job_id: str):
    """
    Get project data from generation result (fallback when database save fails).
    Returns the same format as /api/project/{project_id} but from enriched_data.
    """
    result = progress_manager.get_result(job_id)
    if not result or 'enriched_data' not in result:
        raise HTTPException(status_code=404, detail="Generation data not found")
    
    enriched_data = result['enriched_data']
    
    # Format enriched_data to match frontend ProjectData interface
    from api.services.project_service import format_department_name, normalize_department
    
    # Group crew by department
    departments_map = {}
    crew_list = enriched_data.get('crew_list', [])
    for crew_data in crew_list:
        dept = normalize_department(crew_data.get('department', 'OTHER'))
        dept_name = format_department_name(dept)
        if dept_name not in departments_map:
            departments_map[dept_name] = []
        
        departments_map[dept_name].append({
            "id": str(uuid.uuid4()),  # Generate temporary ID
            "role": crew_data.get('role', 'TBD'),
            "name": crew_data.get('name'),
            "phone": crew_data.get('phone'),
            "email": crew_data.get('email'),
            "callTime": "TBD",
            "location": "Set",
        })
    
    # Format as Department[]
    departments = []
    for dept_name, crew_list in departments_map.items():
        departments.append({
            "name": dept_name,
            "count": len(crew_list),
            "expanded": False,
            "crew": crew_list,
        })
    
    # Sort departments
    dept_order = [
        "PRODUCTION", "CAMERA DEPT", "GRIP & ELECTRIC", "ART DEPT",
        "WARDROBE", "HAIR & MAKEUP", "SOUND", "LOCATIONS",
        "TRANSPORTATION", "CATERING", "POST PRODUCTION", "OTHER"
    ]
    departments.sort(key=lambda d: dept_order.index(d["name"]) if d["name"] in dept_order else 999)
    
    # Format production info
    prod_info = enriched_data.get('production_info', {})
    
    # Format call sheets
    schedule_days = enriched_data.get('schedule_days', [])
    call_sheets = []
    for day in schedule_days:
        call_sheets.append({
            "id": str(uuid.uuid4()),
            "dayNumber": day.get('day_number', 1),
            "shootDate": day.get('date') if day.get('date') != 'TBD' else None,
            "generalCrewCall": day.get('crew_call', 'TBD'),
            "weather": {
                "high": None,
                "low": None,
                "summary": None,
                "sunrise": None,
                "sunset": None,
            },
            "hospital": {
                "name": None,
                "address": None,
            },
        })
    
    # Format locations
    locations_data = enriched_data.get('logistics', {}).get('locations', [])
    locations = []
    for loc in locations_data:
        locations.append({
            "id": str(uuid.uuid4()),
            "name": loc.get('name', 'TBD'),
            "address": loc.get('address'),
            "city": None,
            "state": None,
            "mapLink": loc.get('map_link'),
            "parkingNotes": loc.get('parking'),
        })
    
    return {
        "project": {
            "id": job_id,  # Use job_id as temporary project ID
            "jobName": prod_info.get('job_name', 'TBD'),
            "jobNumber": prod_info.get('job_number', 'TBD'),
            "client": prod_info.get('client', 'TBD'),
            "agency": prod_info.get('agency'),
        },
        "callSheets": call_sheets,
        "locations": locations,
        "departments": departments,
    }


@app.get("/api/project/{project_id}")
async def get_project(project_id: str, db: AsyncSession = Depends(get_db)):
    """
    Get project data formatted for the frontend.

    Returns departments[], callSheets[], locations, and project info
    matching the frontend's TypeScript interfaces.
    """
    data = await get_project_for_frontend(db, project_id)
    if not data:
        raise HTTPException(status_code=404, detail="Project not found")
    return data


@app.patch("/api/project/{project_id}")
async def update_project_endpoint(
    project_id: str,
    jobName: Optional[str] = Query(None, description="New job name"),
    client: Optional[str] = Query(None, description="New client name"),
    agency: Optional[str] = Query(None, description="New agency name"),
    db: AsyncSession = Depends(get_db)
):
    """
    Update project fields (job name, client, agency).
    
    Accepts query parameters for the fields to update.
    """
    success = await update_project(
        db,
        project_id,
        job_name=jobName,
        client=client,
        agency=agency
    )
    if not success:
        raise HTTPException(status_code=404, detail="Project not found")
    
    return {"status": "updated", "project_id": project_id}


@app.post("/api/project/{project_id}/regenerate")
async def regenerate_workbook_endpoint(
    project_id: str,
    db: AsyncSession = Depends(get_db)
):
    """
    Regenerate Excel workbook from current database state.

    Called after data changes (chat edits, crew updates) to sync the Excel file
    with the latest project data.

    Returns:
        - filename: New workbook filename for download
        - download_url: Full URL path to download the file
    """
    # Fetch project data in generator format
    generator_data = await get_project_as_generator_data(db, project_id)
    if not generator_data:
        raise HTTPException(status_code=404, detail="Project not found")

    try:
        # Generate new workbook
        temp_filename = f"temp_regen_{project_id[:8]}.xlsx"
        generator = EpitomeWorkbookGenerator(generator_data, temp_filename)
        generator.generate()

        # Move to output directory with new timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        new_filename = f"Epitome_Workbook_{project_id[:8]}_{timestamp}.xlsx"
        new_path = OUTPUT_DIR / new_filename

        # Move the generated file
        temp_path = Path(temp_filename)
        if temp_path.exists():
            shutil.move(str(temp_path), str(new_path))

        return {
            "filename": new_filename,
            "download_url": f"/api/download/{new_filename}",
        }

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to regenerate workbook: {str(e)}"
        )


@app.patch("/api/crew/{crew_id}")
async def update_crew_member(
    crew_id: str,
    callTime: Optional[str] = Query(None, description="New call time (e.g., '7:00 AM')"),
    location: Optional[str] = Query(None, description="New location (e.g., 'Set', 'Basecamp')"),
    callSheetId: Optional[str] = Query(None, description="Specific call sheet ID to update"),
    db: AsyncSession = Depends(get_db)
):
    """
    Update a crew member's call time or location for a call sheet.

    If callSheetId is provided, updates that specific day.
    Otherwise, updates the first available RSVP.
    """
    success = await update_crew_rsvp(db, crew_id, callTime, location, callSheetId)
    if not success:
        raise HTTPException(status_code=404, detail="Crew member not found")

    return {"status": "updated", "crew_id": crew_id}


@app.get("/api/crew/search")
async def search_crew(
    q: str = Query("", description="Search query (name or role)"),
    department: Optional[str] = Query(None, description="Filter by department"),
    db: AsyncSession = Depends(get_db)
):
    """
    Search available crew members from the organization's database.

    Used by the AddCrewDropdown component to find crew to add to a project.
    """
    crew = await search_crew_members(db, query=q, department=department)
    return {"crew": crew}


@app.post("/api/chat")
async def chat_endpoint(
    project_id: str = Form(...),
    message: str = Form(...),
    history: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_db)
):
    """
    Chat endpoint for asking questions and executing edit commands.

    Accepts:
    - project_id: The project ID to chat about
    - message: User's message/question/command
    - history: JSON string of conversation history (optional)

    Returns:
    - type: "answer" or "edit"
    - response: Text response to show user
    - action: (if edit) The action that was executed
    - success: (if edit) Whether the edit was successful
    """
    # Parse history from JSON string
    history_list = []
    if history:
        try:
            history_list = json.loads(history)
        except json.JSONDecodeError:
            pass  # Ignore malformed history

    result = await process_chat_message(db, project_id, message, history_list)
    return result


@app.get("/health")
async def health_check():
    """Health check endpoint."""
    return {"status": "healthy", "service": "epitome-api"}


# Catch-all route for SPA routing (Vite React Router)
# MUST be last - catches all non-API routes
@app.get("/{full_path:path}")
async def serve_spa(full_path: str):
    """Serve SPA routes - return index.html for non-API routes."""
    # Don't interfere with API routes, health check, or static assets
    # FastAPI will handle /assets/ and other static files via mounted StaticFiles
    if (full_path.startswith("api/") or 
        full_path == "health"):
        raise HTTPException(status_code=404, detail="Not found")
    
    # For any other route (including root paths that don't match mounted routes),
    # serve index.html for SPA routing
    # FastAPI's mounted StaticFiles will handle /assets/ requests before this route
    index_path = STATIC_DIR / "index.html"
    if index_path.exists():
        return FileResponse(str(index_path))
    raise HTTPException(status_code=404, detail="Frontend not found")
