"""
Test extraction with the Fosi level up.xlsx file
"""
import os
import sys
from pathlib import Path

# Load environment variables
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

sys.path.insert(0, str(Path(__file__).parent))

from agents.production_workbook_generator import run_tool

def extract_excel_text(excel_path: str) -> str:
    """Extract text from Excel file."""
    try:
        import openpyxl
        
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        
        extracted_sheets = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_rows = []
            
            for row in sheet.iter_rows(values_only=True):
                # Filter out empty rows
                row_values = [str(cell) if cell is not None else "" for cell in row]
                if any(val.strip() for val in row_values):
                    sheet_rows.append("\t".join(row_values))
            
            if sheet_rows:
                extracted_sheets.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(sheet_rows))
        
        if extracted_sheets:
            return f"[Excel file: {Path(excel_path).name}]\n\n" + "\n\n".join(extracted_sheets)
        else:
            return f"[Excel file: {Path(excel_path).name} - No data found]"
    
    except Exception as e:
        return f"[Excel file: {Path(excel_path).name} - Error extracting data: {str(e)}]"

if __name__ == "__main__":
    excel_path = os.path.expanduser("~/Downloads/Fosi level up.xlsx")
    
    if not os.path.exists(excel_path):
        print(f"❌ File not found: {excel_path}")
        sys.exit(1)
    
    print(f"📄 Extracting Excel file: {excel_path}")
    print(f"📊 File size: {os.path.getsize(excel_path):,} bytes")
    
    excel_text = extract_excel_text(excel_path)
    print(f"✅ Extracted {len(excel_text):,} characters")
    print(f"\n📋 First 500 characters:\n{excel_text[:500]}...")
    
    prompt = "Create call sheets based on the information in this Excel file"
    
    print("\n🤖 Running full extraction pipeline...")
    print("=" * 60)
    
    def progress_callback(stage_id: str, percent: int, message: str):
        print(f"[{percent:3d}%] {stage_id}: {message}")
    
    try:
        result = run_tool(
            prompt=prompt,
            attached_file_content=excel_text,
            enrich=False,  # Skip enrichment to see raw extraction
            progress_callback=progress_callback
        )
        
        print("\n" + "=" * 60)
        print("📊 EXTRACTION RESULTS:")
        print("=" * 60)
        
        data = result['data']
        prod_info = data.get('production_info', {})
        
        print(f"✅ Job Name: {prod_info.get('job_name', 'N/A')}")
        print(f"✅ Job Number: {prod_info.get('job_number', 'N/A')}")
        print(f"✅ Client: {prod_info.get('client', 'N/A')}")
        print(f"✅ Production Company: {prod_info.get('production_company', 'N/A')}")
        print(f"\n📅 Schedule Days: {len(data.get('schedule_days', []))}")
        if data.get('schedule_days'):
            for i, day in enumerate(data['schedule_days'][:5], 1):
                print(f"   Day {day.get('day_number', i)}: {day.get('date', 'N/A')}")
        
        print(f"\n👥 Crew Members: {len(data.get('crew_list', []))}")
        crew_with_data = [c for c in data.get('crew_list', []) if c.get('rate') and c.get('rate') != 'TBD']
        print(f"   Crew with rates: {len(crew_with_data)}")
        if crew_with_data:
            print("   Sample crew with rates:")
            for crew in crew_with_data[:5]:
                print(f"     - {crew.get('role', 'N/A')}: ${crew.get('rate', 'N/A')}")
        
        locations = data.get('logistics', {}).get('locations', [])
        print(f"\n📍 Locations: {len(locations)}")
        if locations:
            for loc in locations[:3]:
                print(f"   - {loc.get('name', 'N/A')}: {loc.get('address', 'N/A')}")
        
        print(f"\n✅ Workbook generated: {result['workbook_path']}")
        print("\n🎉 Test completed successfully!")
        
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
